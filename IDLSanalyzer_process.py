import numpy as np
import sys
import os
import time as tm
import openpyxl as xls
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QListWidgetItem, QLineEdit, QLabel, QRadioButton, QGridLayout, QPushButton, QAction, QActionGroup, QMenu, QInputDialog
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from uiLSProcess import Ui_IDLSanalyzer
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
from IDLSanalyzer_dpss import LSana_dpss
from semiconductor.recombination.intrinsic import Radiative as rad
from semiconductor.recombination.intrinsic import Auger as aug
from semiconductor.material.intrinsic_carrier_density import IntrinsicCarrierDensity as NI
from semiconductor.electrical.ionisation import Ionisation as Ion
import scipy.constants as const
import pandas as pd


class rawdata(object):
    def __init__(self, name=None, Ndop=None, temp=None, doptype=None, nxc=None, tau=None, PCPL='PC', thickness=None, j0=None, plotopt=True, uid=None, **kwarg):
        self.name = name
        self.Ndop = Ndop
        self.temp = temp
        self.doptype = doptype
        self.nxc = nxc
        self.tau = tau
        self.uid = uid
        self.plotopt = plotopt
        self.PCPL = PCPL
        self.thickness = thickness
        self.j0 = j0

    def checkset(self):
        if (self.Ndop is None) or (self.temp is None) or (self.doptype is None):
            return False
        else:
            return True


class LSana(QMainWindow, Ui_IDLSanalyzer):

    def __init__(self, parent=None):
        self.Rawdata = []
        QMainWindow.__init__(self, parent)
        self.setupUi(self)

        self.listWidget_data.setSelectionMode(
            QtWidgets.QAbstractItemView.ExtendedSelection)
        self.listWidget_data.itemSelectionChanged.connect(self.checkselection)
        self.listWidget_data.itemDoubleClicked.connect(self.Opensetting)
        self.listWidget_data.setContextMenuPolicy(Qt.CustomContextMenu)
        self.listWidget_data.customContextMenuRequested.connect(
            self.itemrightclicked)

        self.listWidget_plotlist.setSelectionMode(
            QtWidgets.QAbstractItemView.ExtendedSelection)
        self.listWidget_plotlist.itemSelectionChanged.connect(
            self.checkselection)

        self.listWidget_analysis.setSelectionMode(
            QtWidgets.QAbstractItemView.ExtendedSelection)
        self.listWidget_analysis.itemSelectionChanged.connect(
            self.checkselection)

        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)
        self.verticalLayout.addWidget(self.canvas)
        self.verticalLayout.addWidget(self.toolbar)
        self.ax1 = self.figure.add_subplot(111)
        self.ax1.set_xlabel(r'Excess carrier density $[cm^{-3}]$')
        self.ax1.set_ylabel('Lifetime [s]')
        self.figure.tight_layout()

        self.pushButton_load.clicked.connect(self.loaddata)

        self.pushButton_del.setEnabled(False)
        self.pushButton_del.clicked.connect(self.deldata)

        self.pushButton_data2plot.clicked.connect(self.add2plot)
        self.pushButton_data2plot.setEnabled(False)

        self.pushButton_delplot.clicked.connect(self.delplot)
        self.pushButton_delplot.setEnabled(False)

        self.pushButton_set.clicked.connect(self.Opensetting)
        self.pushButton_set.setEnabled(False)

        self.pushButton_data2fit.clicked.connect(self.add2analysis)
        self.pushButton_data2fit.clicked.connect(self.checkselection)
        self.pushButton_data2fit.setEnabled(False)

        self.pushButton_delanalysis.clicked.connect(self.delAnalysis)
        self.pushButton_delanalysis.clicked.connect(self.checkselection)
        self.pushButton_delanalysis.setEnabled(False)

        self.pushButton_viewset.clicked.connect(self.viewset)
        self.pushButton_viewset.clicked.connect(self.checkselection)
        self.pushButton_viewset.setEnabled(False)

        self.comboBox_plotwhat.currentIndexChanged.connect(self.updatedataplot)
        self.comboBox_plotinverse.currentIndexChanged.connect(
            self.updatedataplot)

        self.pushButton_expdata.clicked.connect(self.export)

        self.pushButton_corp.clicked.connect(self.croppreview)
        self.pushButton_corp.setEnabled(False)

        self.pushButton_intcorrect.clicked.connect(self.correctintrinsic)
        self.pushButton_intcorrect.setEnabled(False)

        self.pushButton_sub2data.clicked.connect(self.sub2)
        self.pushButton_sub2data.setEnabled(False)

        self.pushButton_mergetwo.clicked.connect(self.merge2)
        self.pushButton_mergetwo.setEnabled(False)

        self.pushButton_Analysis.clicked.connect(self.StartAnalysis)
        self.pushButton_Analysis.setEnabled(False)

        self.lineEdit_croplow.textChanged[str].connect(self.checkcorp)
        self.lineEdit_crophigh.textChanged[str].connect(self.checkcorp)
        self.lineEdit_j0.textChanged[str].connect(self.checkj0)
        self.lineEdit_thickness.textChanged[str].connect(self.checkj0)

        self.pushButton_corpconfirm.setEnabled(False)
        self.pushButton_corpconfirm.clicked.connect(self.crop)

        self.pushButton_corpcancel.setEnabled(False)
        self.pushButton_corpcancel.clicked.connect(self.cropcancel)

        self.pushButton_j0correction.clicked.connect(self.j0correction)
        self.pushButton_j0correction.setEnabled(False)

        availablNI = NI().available_models()
        availablRad = rad().available_models()
        availablAug = aug().available_models()
        availablIon = Ion().available_models()

        self.menubar = self.menuBar()
        self.choosmodel = self.menubar.addMenu('Choose your models')
        self.nimodel = self.choosmodel.addMenu('ni models')
        self.Ionmodel = self.choosmodel.addMenu('Ionisation models')
        self.Radmodel = self.choosmodel.addMenu('Radiative models')
        self.Augmodel = self.choosmodel.addMenu('Auger models')

        self.nigroup = QActionGroup(self, exclusive=True)
        self.radgroup = QActionGroup(self, exclusive=True)
        self.auggroup = QActionGroup(self, exclusive=True)
        self.iongroup = QActionGroup(self, exclusive=True)

        for nimodel in availablNI:
            a = self.nigroup.addAction(QAction(nimodel, checkable=True))
            if nimodel == 'Couderc_2014':
                a.setChecked(True)
            self.nimodel.addAction(a)
        for radmodel in availablRad:
            b = self.radgroup.addAction(QAction(radmodel, checkable=True))
            if radmodel == 'Altermatt_2005':
                b.setChecked(True)
            self.Radmodel.addAction(b)
        for augmodel in availablAug:
            c = self.auggroup.addAction(QAction(augmodel, checkable=True))
            if augmodel == 'Richter2012':
                c.setChecked(True)
            self.Augmodel.addAction(c)
        for ionmodel in availablIon:
            d = self.iongroup.addAction(QAction(ionmodel, checkable=True))
            if ionmodel == 'Altermatt_2006_table1':
                d.setChecked(True)
            self.Ionmodel.addAction(d)

    def loaddata(self):
        filename = QFileDialog.getOpenFileNames(
            self, caption='Choose data file', filter="Exported file (*.txt);;Sinton file(*.xlsm)")
        if filename[0] != '':
            for fname in filename[0]:
                if os.path.splitext(fname)[1] == '.txt':
                    data = np.genfromtxt(fname, delimiter='\t', names=[
                                         'Time', 'nxcPC', 'nxcPL', 'tauPC', 'tauPL'])
                    nxcPC = data['nxcPC'][1:]
                    nxcPL = data['nxcPL'][1:]
                    tauPC = data['tauPC'][1:]
                    tauPL = data['tauPL'][1:]
                    idx = np.argsort(nxcPC)
                    nxcPC = nxcPC[idx]
                    tauPC = tauPC[idx]
                    idx = np.argsort(nxcPL)
                    nxcPL = nxcPL[idx]
                    tauPL = tauPL[idx]
                    PL = rawdata(
                        name='PL_' + (os.path.splitext(os.path.basename(fname))[0]), nxc=nxcPL, tau=tauPL, PCPL='PL')
                    PC = rawdata(
                        name='PC_' + (os.path.splitext(os.path.basename(fname))[0]), nxc=nxcPC, tau=tauPC, PCPL='PC')
                    self.Rawdata.append(PL)
                    self.Rawdata.append(PC)
                elif os.path.splitext(fname)[1] == '.xlsm':
                    nxc = []
                    tau = []
                    wb = xls.load_workbook(filename=fname, data_only=True)
                    ws = wb['RawData']
                    for col in ws.iter_cols(min_col=4, max_col=10, min_row=1,max_row=1):
                        if col[0].value == 'Tau (sec)':
                            taucol = col[0].column
                            break
                    for col in ws.iter_cols(min_col=4, max_col=10, min_row=1,max_row=1):
                        if col[0].value == 'Minority Carrier Density':
                            mcdcol = col[0].column
                            break
                    for row in ws.iter_rows(taucol + '5:' + taucol + str(lenth)):
                        if row[0].value is not None and row[-1].value is not None:
                            tau.append(row[0].value)
                    for row in ws.iter_rows(mcdcol + '5:' + mcdcol + str(lenth)):
                        if row[0].value is not None and row[-1].value is not None:
                            nxc.append(row[-1].value)
                    nxc = np.asarray(nxc)
                    tau = np.asarray(tau)
                    idx = np.argsort(nxc)
                    nxc = nxc[idx]
                    tau = tau[idx]
                    ws1 = wb['User']
                    Ndop = ws1['J9'].value
                    doptype = ws1['D6'].value[0]
                    thickness = ws1['B6'].value
                    j0 = ws1['D9'].value
                    if ws1['L8'].value[0] == 'T':
                        temp = ws1['L9'].value + 273.15
                    else:
                        temp = 303
                    Sinton = rawdata(
                        name='Sinton_' + (os.path.splitext(os.path.basename(fname))[0]), nxc=nxc, tau=tau, PCPL='PC', Ndop=Ndop, doptype=doptype, temp=temp, thickness=thickness, j0=j0)
                    self.Rawdata.append(Sinton)

                self.updateuid()
                self.updateDataList()
                self.updateplotlist()
                self.updatedataplot()

    def updateDataList(self):
        self.listWidget_data.clear()
        for data in self.Rawdata:
            dataitem = QListWidgetItem(parent=self.listWidget_data)
            dataitem.setText(data.name)
            dataitem.setData(32, data.uid)
            self.listWidget_data.addItem(dataitem)
            if data.checkset() is False:
                dataitem.setForeground(QtGui.QBrush(QtGui.QColor(255, 0, 0)))
            if data.checkset() is True:
                dataitem.setForeground(QtGui.QBrush(QtGui.QColor(0, 255, 0)))

    def updateplotlist(self):
        self.listWidget_plotlist.clear()
        for data in self.Rawdata:
            if data.plotopt:
                dataitem = QListWidgetItem(parent=self.listWidget_plotlist)
                dataitem.setText(data.name)
                dataitem.setData(32, data.uid)
                self.listWidget_plotlist.addItem(dataitem)

    def updateuid(self):
        for data in self.Rawdata:
            data.uid = self.Rawdata.index(data)

    def updatedataplot(self):
        self.ax1.clear()
        self.ax1.grid()
        self.ax1.set_xlabel(r'Excess carrier density $[cm^{-3}]$')
        if self.comboBox_plotwhat.currentIndex() == 0:
            for data in self.Rawdata:
                if data.plotopt:
                    if self.comboBox_plotinverse.currentIndex() == 0:
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.plot(data.nxc, data.tau, '.', label=data.name)
                    elif self.comboBox_plotinverse.currentIndex() == 1:
                        self.ax1.set_ylabel('Inverse Lifetime [s-1]')
                        self.ax1.plot(data.nxc, 1. / data.tau,
                                      '.', label=data.name)
        elif self.comboBox_plotwhat.currentIndex() == 1:
            for i in range(self.listWidget_analysis.count()):
                item = self.listWidget_analysis.item(i)
                for data in self.Rawdata:
                    if data.uid == item.data(32):
                        if self.comboBox_plotinverse.currentIndex() == 0:
                            self.ax1.set_ylabel('Lifetime [s]')
                            self.ax1.plot(data.nxc, data.tau,
                                          '.', label=data.name)
                        elif self.comboBox_plotinverse.currentIndex() == 1:
                            self.ax1.set_ylabel('Inverse Lifetime [s-1]')
                            self.ax1.plot(data.nxc, 1. / data.tau,
                                          '.', label=data.name)
        self.ax1.semilogx()
        self.ax1.legend(loc=0)
        self.figure.tight_layout()
        self.canvas.draw()

    def deldata(self):
        for item in self.listWidget_data.selectedItems():
            for data in self.Rawdata:
                if data.uid == item.data(32):
                    self.Rawdata.remove(data)
        self.updateuid()
        self.updateDataList()
        self.updateplotlist()
        self.updatedataplot()

    def delplot(self):
        for item in self.listWidget_plotlist.selectedItems():
            for data in self.Rawdata:
                if data.uid == item.data(32):
                    data.plotopt = False
        self.updateplotlist()
        self.updatedataplot()

    def add2plot(self):
        for item in self.listWidget_data.selectedItems():
            for data in self.Rawdata:
                if data.uid == item.data(32):
                    data.plotopt = True
        self.updateplotlist()
        self.updatedataplot()

    def Opensetting(self):
        self.dialog = QtWidgets.QDialog()
        grid = QGridLayout(self.dialog)
        self.ntype = QRadioButton('n-type')
        self.ptype = QRadioButton('p-type')
        self.dop = QLineEdit()
        self.temp = QLineEdit()
        self.jj0 = QLineEdit()
        self.thick = QLineEdit()
        Ldop = QLabel('Ndop (cm-3)')
        Ltemp = QLabel('Temp (K)')
        Lj0 = QLabel('J0 (A/cm2)')
        Lthick = QLabel('Thickness (cm)')
        self.ok = QPushButton('OK')
        self.ok.setEnabled(False)
        grid.addWidget(self.ntype, 0, 0)
        grid.addWidget(self.ptype, 0, 1)
        grid.addWidget(Ldop, 1, 0)
        grid.addWidget(self.dop, 1, 1)
        grid.addWidget(Ltemp, 2, 0)
        grid.addWidget(self.temp, 2, 1)
        grid.addWidget(Lj0, 3, 0)
        grid.addWidget(self.jj0, 3, 1)
        grid.addWidget(Lthick, 4, 0)
        grid.addWidget(self.thick, 4, 1)
        grid.addWidget(self.ok, 5, 1)
        for data in self.Rawdata:
            if data.uid == self.listWidget_data.selectedItems()[0].data(32):
                if data.Ndop is not None:
                    self.dop.setText('{:e}'.format(data.Ndop))
                if data.temp is not None:
                    self.temp.setText(str(data.temp))
                if data.doptype == 'n':
                    self.ntype.setChecked(True)
                if data.doptype == 'p':
                    self.ptype.setChecked(True)
                if data.j0 is not None:
                    self.jj0.setText('{:e}'.format(data.j0))
                if data.thickness is not None:
                    self.thick.setText(str(data.thickness))
        self.dop.textChanged[str].connect(self.checkparam)
        self.temp.textChanged[str].connect(self.checkparam)
        self.thick.textChanged[str].connect(self.checkparam)
        self.jj0.textChanged[str].connect(self.checkparam)
        self.ntype.toggled.connect(self.checkparam)
        self.ok.clicked.connect(self.setparam)
        self.dialog.exec_()

    def checkparam(self):
        try:
            if float(self.dop.text()) > 0 and float(self.temp.text()) > 0 and len(self.listWidget_data.selectedItems()) == 1:
                self.ok.setEnabled(True)
            else:
                self.ok.setEnabled(False)
        except ValueError:
            self.ok.setEnabled(False)
        if self.thick.text() != "":
            try:
                if float(self.thick.text()) < 0:
                    self.ok.setEnabled(False)
            except ValueError:
                self.ok.setEnabled(False)
        if self.jj0.text() != "":
            try:
                if float(self.jj0.text()) < 0:
                    self.ok.setEnabled(False)
            except ValueError:
                self.ok.setEnabled(False)

    def checkcorp(self):
        self.pushButton_corpconfirm.setEnabled(False)
        self.pushButton_corpcancel.setEnabled(False)
        try:
            if float(self.lineEdit_croplow.text()) >= float(self.lineEdit_crophigh.text()):
                self.pushButton_corp.setEnabled(False)
            elif len(self.listWidget_data.selectedItems()) > 0:
                self.pushButton_corp.setEnabled(True)
        except:
            self.pushButton_corp.setEnabled(False)

    def setparam(self):
        for data in self.Rawdata:
            if data.uid == self.listWidget_data.selectedItems()[0].data(32):
                data.Ndop = float(self.dop.text())
                data.temp = float(self.temp.text())
                try:
                    data.j0 = float(self.jj0.text())
                    data.thickness = float(self.thick.text())
                except ValueError:
                    pass
                if self.ntype.isChecked():
                    data.doptype = 'n'
                if self.ptype.isChecked():
                    data.doptype = 'p'
        self.updateDataList()
        self.dialog.close()

    def checkselection(self):
        if len(self.listWidget_data.selectedItems()) > 0:
            self.pushButton_del.setEnabled(True)
            self.pushButton_corp.setEnabled(True)
            self.pushButton_data2plot.setEnabled(True)
            i = 0
            for item in self.listWidget_data.selectedItems():
                for data in self.Rawdata:
                    if data.uid == item.data(32):
                        if data.checkset() is True:
                            i += 1
            if i == 0:
                self.pushButton_data2fit.setEnabled(False)
                self.pushButton_intcorrect.setEnabled(False)
                self.pushButton_j0correction.setEnabled(False)
            elif i == 1:
                self.pushButton_j0correction.setEnabled(True)
                self.pushButton_data2fit.setEnabled(True)
                self.pushButton_intcorrect.setEnabled(True)
            else:
                self.pushButton_data2fit.setEnabled(True)
                self.pushButton_intcorrect.setEnabled(True)
                self.pushButton_j0correction.setEnabled(False)

            if len(self.listWidget_data.selectedItems()) == 1:
                self.pushButton_set.setEnabled(True)
                for item in self.listWidget_data.selectedItems():
                    for data in self.Rawdata:
                        if data.uid == item.data(32):
                            if data.j0 is not None:
                                self.lineEdit_j0.setText(
                                    '{:.2e}'.format(data.j0))
                            else:
                                self.lineEdit_j0.clear()
                            if data.thickness is not None:
                                self.lineEdit_thickness.setText(
                                    str(data.thickness))
                            else:
                                self.lineEdit_thickness.clear()
            else:
                self.pushButton_set.setEnabled(False)
                self.pushButton_intcorrect.setEnabled(False)
                self.lineEdit_j0.clear()
                self.lineEdit_thickness.clear()
            if len(self.listWidget_data.selectedItems()) == 2:
                self.pushButton_sub2data.setEnabled(True)
                self.pushButton_mergetwo.setEnabled(True)
            else:
                self.pushButton_sub2data.setEnabled(False)
                self.pushButton_mergetwo.setEnabled(False)
        else:
            self.pushButton_sub2data.setEnabled(False)
            self.pushButton_mergetwo.setEnabled(False)
            self.pushButton_intcorrect.setEnabled(False)
            self.pushButton_corp.setEnabled(False)
            self.pushButton_del.setEnabled(False)
            self.pushButton_data2plot.setEnabled(False)
            self.pushButton_set.setEnabled(False)
            self.pushButton_data2fit.setEnabled(False)
            self.lineEdit_j0.clear()
            self.lineEdit_thickness.clear()

        if len(self.listWidget_analysis.selectedItems()) > 0:
            self.pushButton_delanalysis.setEnabled(True)
            self.pushButton_viewset.setEnabled(True)
        else:
            self.pushButton_delanalysis.setEnabled(False)
            self.pushButton_viewset.setEnabled(False)

        if len(self.listWidget_plotlist.selectedItems()) > 0:
            self.pushButton_delplot.setEnabled(True)
        else:
            self.pushButton_delplot.setEnabled(False)

        if self.listWidget_analysis.count() > 0:
            self.pushButton_Analysis.setEnabled(True)
        else:
            self.pushButton_Analysis.setEnabled(False)
        self.checkcorp()
        self.checkj0()

    def checkj0(self):
        self.pushButton_j0correction.setEnabled(False)
        try:
            if float(self.lineEdit_j0.text()) > 0 and float(self.lineEdit_thickness.text()) > 0:
                self.pushButton_j0correction.setEnabled(True)
            else:
                self.pushButton_j0correction.setEnabled(False)
        except:
            self.pushButton_j0correction.setEnabled(False)

    def add2analysis(self):
        for item in self.listWidget_data.selectedItems():
            for data in self.Rawdata:
                if data.uid == item.data(32) and data.uid not in self.getnalysislist():
                    if data.checkset() == True:
                        dataitem = QListWidgetItem(
                            parent=self.listWidget_analysis)
                        dataitem.setText(data.name)
                        dataitem.setData(32, data.uid)
                        self.listWidget_analysis.addItem(dataitem)

    def getnalysislist(self):
        uidlist = []
        for i in range(self.listWidget_analysis.count()):
            uidlist.append(self.listWidget_analysis.item(i).data(32))
        return uidlist

    def delAnalysis(self):
        for item in self.listWidget_analysis.selectedItems():
            self.listWidget_analysis.takeItem(
                self.listWidget_analysis.row(item))

    def export(self):
        path = QFileDialog.getExistingDirectory(
            parent=self, caption='Select path for data export')
        fname = os.path.join(
            path, str(tm.mktime(tm.localtime())) + '_lifetimeplot.csv')
        data = {}
        for line in self.ax1.get_lines():
            data[line.get_label() + '_' + self.ax1.get_xlabel()
                 ] = line.get_xdata()
            data[line.get_label() + '_' + self.ax1.get_ylabel()
                 ] = line.get_ydata()
        df = pd.DataFrame.from_dict(
            data, orient='index').transpose().fillna('')
        df.to_csv(fname, index=False)

    def croppreview(self):
        self.ax1.clear()
        self.ax1.grid()
        self.ax1.set_xlabel(r'Excess carrier density $[cm^{-3}]$')
        for item in self.listWidget_data.selectedItems():
            for data in self.Rawdata:
                if data.uid == item.data(32):
                    index = data.nxc >= float(self.lineEdit_croplow.text())
                    index *= data.nxc <= float(self.lineEdit_crophigh.text())
                    tau = data.tau[index]
                    nxc = data.nxc[index]
                    if self.comboBox_plotinverse.currentIndex() == 0:
                        self.ax1.set_ylabel('Lifetime [s]')
                        self.ax1.plot(data.nxc, data.tau, '.', label=data.name)
                        self.ax1.plot(nxc, tau, 'o', label='crop_' + data.name)
                    elif self.comboBox_plotinverse.currentIndex() == 1:
                        self.ax1.set_ylabel('Inverse Lifetime [s-1]')
                        self.ax1.plot(data.nxc, 1. / data.tau,
                                      '.', label=data.name)
                        self.ax1.plot(nxc, 1. / tau, 'o',
                                      label='crop_' + data.name)
        self.ax1.semilogx()
        self.ax1.legend(loc=0)
        self.canvas.draw()
        self.figure.tight_layout()
        self.pushButton_corpconfirm.setEnabled(True)
        self.pushButton_corpcancel.setEnabled(True)
        self.comboBox_plotwhat.setEnabled(False)
        self.comboBox_plotinverse.setEnabled(False)
        self.listWidget_plotlist.setEnabled(False)
        self.pushButton_delplot.setEnabled(False)
        self.pushButton_intcorrect.setEnabled(False)
        self.pushButton_sub2data.setEnabled(False)
        self.pushButton_mergetwo.setEnabled(False)
        self.groupBox.setEnabled(False)
        self.groupBox_2.setEnabled(False)

    def crop(self):
        for item in self.listWidget_data.selectedItems():
            crop = rawdata()
            for data in self.Rawdata:
                if data.uid == item.data(32):
                    data.plotopt = False
                    index = data.nxc >= float(self.lineEdit_croplow.text())
                    index *= data.nxc <= float(self.lineEdit_crophigh.text())
                    tau = data.tau[index]
                    nxc = data.nxc[index]
                    crop.name = 'crop_' + data.name
                    crop.Ndop = data.Ndop
                    crop.temp = data.temp
                    crop.doptype = data.doptype
                    crop.nxc = nxc
                    crop.tau = tau
                    crop.PCPL = data.PCPL
            self.Rawdata.append(crop)

        self.updateuid()
        self.updateDataList()
        self.updateplotlist()
        self.cropcancel()

    def cropcancel(self):
        self.updatedataplot()
        self.pushButton_corpconfirm.setEnabled(False)
        self.pushButton_corpcancel.setEnabled(False)
        self.comboBox_plotwhat.setEnabled(True)
        self.comboBox_plotinverse.setEnabled(True)
        self.listWidget_plotlist.setEnabled(True)
        self.groupBox.setEnabled(True)
        self.groupBox_2.setEnabled(True)
        self.checkselection()

    def correctintrinsic(self):
        for action in self.nimodel.actions():
            if action.isChecked():
                ni_author = action.text()
        for action in self.Radmodel.actions():
            if action.isChecked():
                radauthor = action.text()
        for action in self.Augmodel.actions():
            if action.isChecked():
                augauthor = action.text()
        for action in self.Ionmodel.actions():
            if action.isChecked():
                ionauthor = action.text()
        for item in self.listWidget_data.selectedItems():
            intrinsic = rawdata()
            tauint = rawdata()
            for data in self.Rawdata:
                if data.uid == item.data(32) and data.checkset() == True:
                    intrinsic.name = 'IntrinsicCorrected_' + data.name
                    intrinsic.Ndop = data.Ndop
                    intrinsic.temp = data.temp
                    intrinsic.doptype = data.doptype
                    intrinsic.PCPL = data.PCPL
                    intrinsic.nxc = data.nxc
                    tauint.name = 'Intrinsic_' + data.name
                    tauint.Ndop = data.Ndop
                    tauint.temp = data.temp
                    tauint.doptype = data.doptype
                    tauint.PCPL = data.PCPL
                    tauint.nxc = data.nxc
                    ni = NI(temp=data.temp).update(author=ni_author)
                    if data.doptype == 'n':
                        Nd = Ion(temp=data.temp, ni_author=ni_author).update_dopant_ionisation(
                            author=ionauthor, N_dop=data.Ndop, nxc=0, impurity='phosphorous')
                        Na = 1
                    if data.doptype == 'p':
                        Na = Ion(temp=data.temp, ni_author=ni_author).update_dopant_ionisation(
                            author=ionauthor, N_dop=data.Ndop, nxc=0, impurity='boron')
                        Nd = 1
                    itauintrin = rad().itau(ni_author=ni_author, author=radauthor, temp=data.temp, Na=Na, Nd=Nd, nxc=data.nxc) + \
                        aug().itau(ni_author=ni_author, author=augauthor,
                                   temp=data.temp, Na=Na, Nd=Nd, nxc=data.nxc)
                    intrinsic.tau = 1. / (1. / data.tau - itauintrin)
                    tauint.tau = 1. / (itauintrin)
            self.Rawdata.append(intrinsic)
            self.Rawdata.append(tauint)
        self.updateuid()
        self.updateDataList()
        self.updateplotlist()
        self.updatedataplot()

    def sub2(self):
        sub = rawdata()
        item1 = self.listWidget_data.selectedItems()[0]
        item2 = self.listWidget_data.selectedItems()[1]
        for data in self.Rawdata:
            if data.uid == item1.data(32):
                sub1 = data
                data.plotopt = False
            if data.uid == item2.data(32):
                sub2 = data
                data.plotopt = False
        if sub1.Ndop == sub2.Ndop:
            sub.Ndop = sub1.Ndop
        if sub1.temp == sub2.temp:
            sub.temp = sub1.temp
        if sub1.doptype == sub2.doptype:
            sub.doptype = sub1.doptype
        if sub1.PCPL == sub2.PCPL:
            sub.PCPL = sub1.PCPL
        else:
            sub.PCPL = None
        lb = np.maximum(np.amin(sub1.nxc), np.amin(sub2.nxc))
        ub = np.minimum(np.amax(sub1.nxc), np.amax(sub2.nxc))
        npoints = np.minimum(sub1.nxc[(sub1.nxc >= lb) & (sub1.nxc <= ub)].size, sub2.nxc[(
            sub2.nxc >= lb) & (sub2.nxc <= ub)].size)
        nxc = np.logspace(np.log10(lb), np.log10(ub), npoints)
        tau1 = np.interp(nxc, sub1.nxc, sub1.tau)
        tau2 = np.interp(nxc, sub2.nxc, sub2.tau)
        if np.average(sub1.tau) > np.average(sub2.tau):
            tau = 1. / (1. / tau2 - 1. / tau1)
        elif np.average(sub1.tau) < np.average(sub2.tau):
            tau = 1. / (1. / tau1 - 1. / tau2)
        sub.nxc = nxc
        sub.tau = tau
        sub.name = 'sub_' + sub1.name + '_and_' + sub2.name
        self.Rawdata.append(sub)
        self.updateuid()
        self.updateDataList()
        self.updateplotlist()
        self.updatedataplot()

    def merge2(self):
        merge = rawdata()
        item1 = self.listWidget_data.selectedItems()[0]
        item2 = self.listWidget_data.selectedItems()[1]
        for data in self.Rawdata:
            if data.uid == item1.data(32):
                merge1 = data
                data.plotopt = False
            if data.uid == item2.data(32):
                merge2 = data
                data.plotopt = False
        if merge1.Ndop == merge2.Ndop:
            merge.Ndop = merge1.Ndop
        if merge1.temp == merge2.temp:
            merge.temp = merge1.temp
        if merge1.doptype == merge2.doptype:
            merge.doptype = merge1.doptype
        if merge1.PCPL == merge2.PCPL:
            merge.PCPL = merge1.PCPL
        else:
            merge.PCPL = None
        tau = np.append(merge1.tau, merge2.tau)
        nxc = np.append(merge1.nxc, merge2.nxc)
        idx = np.argsort(nxc)
        nxc = nxc[idx]
        tau = tau[idx]
        merge.nxc = nxc
        merge.tau = tau
        merge.name = 'merge_' + merge1.name + '_and_' + merge2.name
        self.Rawdata.append(merge)
        self.updateuid()
        self.updateDataList()
        self.updateplotlist()
        self.updatedataplot()

    def viewset(self):
        self.dialog2 = QtWidgets.QDialog()
        grid = QGridLayout(self.dialog2)
        Ltype = QLabel('Doping type')
        Type = QLineEdit()
        dop = QLineEdit()
        temp = QLineEdit()
        Ldop = QLabel('Ndop (cm-3)')
        Ltemp = QLabel('Temp (K)')
        grid.addWidget(Ltype, 0, 0)
        grid.addWidget(Type, 0, 1)
        grid.addWidget(Ldop, 1, 0)
        grid.addWidget(dop, 1, 1)
        grid.addWidget(Ltemp, 2, 0)
        grid.addWidget(temp, 2, 1)
        Type.setReadOnly(True)
        dop.setReadOnly(True)
        temp.setReadOnly(True)
        for data in self.Rawdata:
            if data.uid == self.listWidget_analysis.selectedItems()[0].data(32):
                if data.Ndop is not None:
                    dop.setText('{:e}'.format(data.Ndop))
                if data.temp is not None:
                    temp.setText(str(data.temp))
                if data.doptype == 'n':
                    Type.setText('n')
                if data.doptype == 'p':
                    Type.setText('p')
        self.dialog2.exec_()

    def itemrightclicked(self):
        if len(self.listWidget_data.selectedItems()) == 1:
            menu = QMenu(self.listWidget_data)
            Rset = menu.addAction('Set')
            Rset.triggered.connect(self.Opensetting)
            Rcopy = menu.addAction('Copy')
            Rcopy.triggered.connect(self.copydata)
            Rrename = menu.addAction('Rename')
            Rrename.triggered.connect(self.changename)
            Rdel = menu.addAction('Delete')
            Rdel.triggered.connect(self.deldata)
            menu.popup(QtGui.QCursor.pos())

    def copydata(self):
        if len(self.listWidget_data.selectedItems()) == 1:
            item = self.listWidget_data.selectedItems()[0]
            copy = rawdata()
            for data in self.Rawdata:
                if data.uid == item.data(32):
                    copy.name = data.name
                    copy.Ndop = data.Ndop
                    copy.temp = data.temp
                    copy.doptype = data.doptype
                    copy.nxc = data.nxc
                    copy.tau = data.tau
                    copy.PCPL = data.PCPL
            self.Rawdata.append(copy)
            self.updateuid()
            self.updateDataList()
            self.updateplotlist()
            self.updatedataplot()

    def changename(self):
        if len(self.listWidget_data.selectedItems()) == 1:
            item = self.listWidget_data.selectedItems()[0]
            text, ok = QInputDialog.getText(
                self, 'Rename', 'Enter a new name:')
            if ok:
                item.setText(text)
                for data in self.Rawdata:
                    if data.uid == item.data(32):
                        data.name = text
        self.updateDataList()
        self.updateplotlist()
        self.updatedataplot()

    def StartAnalysis(self):
        uidlist = self.getnalysislist()
        self.IDLSanalyzer_dpss = LSana_dpss(
            rawdata=self.Rawdata, uidlist=uidlist, mainwindow=LSana)
        LSana.hide()
        self.IDLSanalyzer_dpss.show()

    def j0correction(self):
        for action in self.Ionmodel.actions():
            if action.isChecked():
                ionauthor = action.text()
        for action in self.nimodel.actions():
            if action.isChecked():
                ni_author = action.text()
        for item in self.listWidget_data.selectedItems():
            J0 = rawdata()
            J0corrected = rawdata()
            for data in self.Rawdata:
                if data.uid == item.data(32) and data.checkset() == True:
                    J0corrected.name = 'J0Corrected_' + data.name
                    J0corrected.Ndop = data.Ndop
                    J0corrected.temp = data.temp
                    J0corrected.doptype = data.doptype
                    J0corrected.PCPL = data.PCPL
                    J0corrected.nxc = data.nxc
                    J0.name = 'J0_' + data.name
                    J0.Ndop = data.Ndop
                    J0.temp = data.temp
                    J0.doptype = data.doptype
                    J0.PCPL = data.PCPL
                    J0.nxc = data.nxc
                    ni = NI(temp=data.temp).update(author=ni_author)
                    if data.doptype == 'n':
                        iNdop = Ion(temp=data.temp, ni_author=ni_author).update_dopant_ionisation(
                            author=ionauthor, N_dop=data.Ndop, nxc=0, impurity='phosphorous')
                    if data.doptype == 'p':
                        iNdop = Ion(temp=data.temp, ni_author=ni_author).update_dopant_ionisation(
                            author=ionauthor, N_dop=data.Ndop, nxc=0, impurity='boron')
                    itauj0 = 2 * float(self.lineEdit_j0.text()) * (iNdop + data.nxc) / \
                        const.e / float(self.lineEdit_thickness.text()) / ni**2
                    J0corrected.tau = 1. / (1. / data.tau - itauj0)
                    J0.tau = 1. / (itauj0)
            self.Rawdata.append(J0corrected)
            self.Rawdata.append(J0)
        self.updateuid()
        self.updateDataList()
        self.updateplotlist()
        self.updatedataplot()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    LSana = LSana()
    LSana.show()
    sys.exit(app.exec_())
